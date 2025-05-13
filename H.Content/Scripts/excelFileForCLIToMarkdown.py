import openpyxl

# This script reads from 

# Set these parameters to your own values
xlsxFilePath = r"C:\Users\ExampleUser\Documents\CLI_Parameters.xlsx"
outputMarkdownFileName = "OutputMarkdown.md"

workbook = openpyxl.load_workbook(xlsxFilePath)

# Create a markdown file or write to existing file
with open(outputMarkdownFileName, "a") as f:

    # Iterate through sheets in xlsx file
    for sheet in workbook.worksheets:

        # Get last populated row, openpyxl does not have a method that does this correctly
        lastPopulatedRow = 0
        for row in range(sheet.max_row, 0, -1):
            if any(cell.value is not None for cell in sheet[row]):
                lastPopulatedRow = row
                break

        # Write the header of the markdown file
        f.write("\n<br>\n\n")
        f.write("# " + sheet.title + "\n")

        for i in range (1, lastPopulatedRow+1):
            for j in range (1, sheet.max_column+1):
                
                # Removes zeros appended to integers when read from xlsx file
                value = (str(sheet.cell(i,j).value) + " ").replace(".0 ", "")
                # Replaces "None" with empty string
                if (value == "None "):
                    value = ""
                
                if (i == 1):
                    break
                elif (j == 1):
                    f.write("## " + str(sheet.cell(i,j).value).replace(" ", "") + "\n\n")
                elif (j == 2):
                    f.write(str(sheet.cell(1,j+1).value) + ": " + value + "\n\n")
                elif (j == 3):
                    f.write(str(sheet.cell(1,j+1).value) + ": " + value + "\n\n")
                elif (j == 4):
                    f.write(str(sheet.cell(1,j+1).value) + ": " + value + "\n\n")
                elif (j == 5):
                    f.write(str(sheet.cell(1,j+1).value) + ": " + value + "\n\n")
                elif (j == 6):
                    f.write(str(sheet.cell(1,j+1).value) + ": " + value + "\n\n")
                elif (j == 7):
                    f.write(str(sheet.cell(1,j+1).value) + ": " + value + "\n\n")
                    f.write("***\n")

                    
                
    